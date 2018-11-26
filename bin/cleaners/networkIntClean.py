###############################################################
## networkInterfacesClean.py
## last updated 25.11.2018
## purpose : networkInterface cleaner module for log parser module
## authors: luckyducky + lamehacker + calebpitts
################################################################

import csv
import gc
import os
import re
from datetime import datetime
from openpyxl import load_workbook

import pandas as pd

LogID = []
captureTime = []
name = []
active= []
speed= []
duplex = []
mtu = []
incomingBytes = []
incomingPackets = []
incomingErrors = []
incomingDrops = []
outgoingBytes = []
outgoingPackets = []
outgoingErrors = []
outgoingDrops = []
IPv4= []
IPv4_netmask = []
IPv4_broadcast=[]
IPv4_p2p=[]
IPv6 = []
IPv6_netmask = []
IPv6_broadcast=[]
IPv6_p2p=[]

def addHeaders():
    global LogID, captureTime, name, active
    global speed, duplex, mtu, incomingBytes
    global incomingPackets, incomingErrors, incomingDrops
    global outgoingBytes, outgoingPackets, outgoingErrors
    global outgoingDrops, IPv4, IPv6, IPv4_netmask, IPv4_netmask
    global IPv4_broadcast, IPv6, IPv4_netmask, IPv4_broadcast
    global IPv4_p2p, IPv6_netmask, IPv6_p2p

    LogID.append('Log ID')
    captureTime.append('Log time')
    name.append('Interface name')
    active.append('Activity detected')
    speed.append('Speed')
    duplex.append('Duplex info')
    mtu.append('MTU')
    incomingBytes.append('Incoming Bytes')
    incomingPackets.append('Incoming Packets')
    incomingErrors.append('Incoming Errors')
    incomingDrops.append('Incoming Drops')
    outgoingBytes.append('Outgoing Bytes')
    outgoingPackets.append('Outgoing Packets')
    outgoingErrors.append('Outgoing Errors')
    outgoingDrops.append('Outgoing Drops')
    IPv4.append('IPv4')
    IPv6.append('IPv6')
    IPv4_netmask.append('IPv4 Netmask')
    IPv4_broadcast.append('IPv4 Broadcast')
    IPv4_p2p.append('IPv4 p2p')
    IPv6_netmask.append('IPv4 netmask')
    IPv6_broadcast.append('IPv6 Broadcast')
    IPv6_p2p.append('IPv6 p2p')

def emptyLists():
    global LogID, captureTime, name, active
    global speed, duplex, mtu, incomingBytes
    global incomingPackets, incomingErrors, incomingDrops
    global outgoingBytes, outgoingPackets, outgoingErrors
    global outgoingDrops, IPv4, IPv6, IPv4_netmask, IPv4_netmask
    global IPv4_broadcast, IPv6, IPv4_netmask, IPv4_broadcast
    global IPv4_p2p, IPv6_netmask, IPv6_p2p

    LogID[:] = []
    captureTime[:] = []
    name[:] = []
    active[:]= []
    speed[:]= []
    duplex[:] = []
    mtu[:] = []
    incomingBytes[:] = []
    incomingPackets[:] = []
    incomingErrors[:] = []
    incomingDrops[:] = []
    outgoingBytes[:] = []
    outgoingPackets[:]= []
    outgoingErrors[:] = []
    outgoingDrops[:] = []
    IPv4[:]= []
    IPv4_netmask[:] = []
    IPv4_broadcast[:]=[]
    IPv4_p2p[:]=[]
    IPv6[:] = []
    IPv6_netmask[:] = []
    IPv6_broadcast[:]=[]
    IPv6_p2p[:]=[]

# function to define structure of LogID
def generateLogID(logCount, timeForLogID, machineID):
    if logCount <= 9 and logCount >= 0:
        placeholder = '000'
    elif logCount <= 99 and logCount >= 10:
        placeholder = '00'
    elif logCount > 99:
        placeholder = '0'
    date = timeForLogID[13:23]
    LogID.append(machineID + '_' + str(date) + '[' + placeholder + str(logCount) + ']')


def initiator(dataDir, machineID, logPath):
    timestampFix = 0
    timeStamp = ''

    logCount = 0
    netmaskIPv4 = 0
    netmaskIPv6 = 0
    broadcastIPv4 = 0
    broadcastIPv6 = 0
    p2pIPv4 = 0
    p2pIPv6 = 0
    ipv4 = 0
    ipv6 = 0

    os.chdir(dataDir)
    if os.path.exists('networkIntLogs.xlsx') == False:
        addHeaders()
    else:
        emptyLists()

    os.chdir(logPath)
    with open('networkInterfaces.log') as networkLog:
        for line in networkLog:

            if line.startswith('</log>') == True:
                logCount += 1

            if line.startswith('<MAC>') == True:
                continue

            if line.startswith('<captureTime>') == True:
                timeStamp = line

            if line.startswith('<name>') == True:
                name.append(line)
                captureTime.append(timeStamp)
                timestampFix += 1
                generateLogID(logCount, timeStamp, machineID)

            elif line.startswith('<active>') == True:
                active.append(line)

            elif line.startswith('<speed>') == True:
                speed.append(line)

            elif line.startswith('<duplex>') == True:
                duplex.append(line)

            elif line.startswith('<mtu>') == True:
                mtu.append(line)

            elif line.startswith('<incomingBytes>') == True:
                incomingBytes.append(line)

            elif line.startswith('<incomingPackets>') == True:
                incomingPackets.append(line)

            elif line.startswith('<incomingErrors>') == True:
                incomingErrors.append(line)

            elif line.startswith('<incomingDrops>') == True:
                incomingDrops.append(line)

            elif line.startswith('<outgoingBytes>') == True:
                outgoingBytes.append(line)

            elif line.startswith('<outgoingPackets>') == True:
                outgoingPackets.append(line)

            elif line.startswith('<outgoingErrors>') == True:
                outgoingErrors.append(line)

            elif line.startswith('<outgoingErrors>') == True:
                outgoingErrors.append(line)

            elif line.startswith('<outgoingDrops>') == True:
                outgoingDrops.append(line)

            elif line.startswith('<IPv4>') == True:
                ipv4 = 1
                IPv4.append(line)

            elif line.startswith('<IPv4_netmask>') == True:
                netmaskIPv4 = 1
                IPv4_netmask.append(line)

            elif line.startswith('<IPv4_broadcast>') == True:
                broadcastIPv4 = 1
                IPv4_broadcast.append(line)

            elif line.startswith('<IPv4_p2p>') == True:
                p2pIPv4 = 1
                IPv4_p2p.append(line)

            elif line.startswith('<IPv6>') == True:
                ipv6 = 1
                IPv6.append(line)

            elif line.startswith('<IPv6_netmask>') == True:
                netmaskIPv6 = 1
                IPv6_netmask.append(line)

            elif line.startswith('<IPv6_broadcast>') == True:
                broadcastIPv6 = 1
                IPv6_broadcast.append(line)

            elif line.startswith('<IPv4_p2p>') == True:
                p2pIPv6 = 1
                IPv4_p2p.append(line)

            if line.startswith('</log>') == True:
                if ipv4 == 0:
                    IPv4.append('')
                if ipv6 == 0:
                    IPv6.append('')
                if netmaskIPv4 == 0:
                    IPv4_netmask.append('')
                if broadcastIPv4 == 0:
                    IPv4_broadcast.append('')
                if netmaskIPv6 == 0:
                    IPv6_netmask.append('')
                if broadcastIPv6 == 0:
                    IPv6_broadcast.append('')
                if p2pIPv4 == 0:
                    IPv4_p2p.append('')
                if p2pIPv6 == 0:
                    IPv6_p2p.append('')

    df = pd.DataFrame(list(zip(LogID, captureTime, name, active, speed, duplex, mtu, incomingBytes, incomingPackets,
                    incomingErrors, incomingDrops, outgoingBytes, outgoingPackets, outgoingErrors, outgoingDrops, IPv4,
                    IPv4_netmask,IPv4_broadcast, IPv4_p2p, IPv6, IPv6_netmask, IPv6_broadcast, IPv6_p2p)))

    stuffToIgnore = ['<captureTime>', '</captureTime>', '<name>', '</name>', '<speed>', '</speed>', '<active>',
                     '</active>', '<duplex>', '</duplex>', '<mtu>', '</mtu>', '<incomingBytes>', '</incomingBytes>',
                     '<incomingPackets>', '</incomingPackets>', '<incomingErrors>', '</incomingErrors>',
                     '<incomingDrops>', '</incomingDrops>', '<outgoingBytes>', '</outgoingBytes>', '<outgoingPackets>',
                     '</outgoingPackets>', '<outgoingErrors>', '</outgoingErrors>', '<outgoingDrops>',
                     '</outgoingDrops>', '<IPv4>', '</IPv4>', '<IPv4_netmask>', '</IPv4_netmask>', '<IPv6>', '</IPv6>',
                     '<IPv6_netmask>', '<MAC>', '</MAC>', '<IPv4_broadcast>', '</IPv4_broadcast>', '<IPv4_p2p>',
                     '</IPv4_p2p>', '<IPv6_p2p>', '</IPv6_p2p>']

    # needed to add a separate line for \n removal cause it wasn't working in the stuffToIgnore
    df = df.replace("\n", "", regex=True)
    for stuff in stuffToIgnore:
        df = df.replace(stuff, "", regex=True)

    os.chdir(dataDir)

    if os.path.exists('networkIntLogs.xlsx') == True:
        oldFile = load_workbook('networkIntLogs.xlsx')
        writer = pd.ExcelWriter('networkInteLogs.xlsx', engine='openpyxl')
        writer.book = oldFile
        writer.sheets = {ws.title: ws for ws in oldFile.worksheets}
        df.to_excel(writer, sheet_name='Network interface logs', startrow=writer.sheets['Network interface logs'].max_row, index=False,
                    header=False)
        writer.save()
    else:
        writer = pd.ExcelWriter('networkLogs.xlsx')
        df.to_excel(writer, sheet_name='networkInterface logs', index=False, header=False)
        writer.save()