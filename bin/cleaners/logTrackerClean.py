###############################################################
# logTrackerClean.py
# last updated 24.11.2018
# purpose : logTracker cleaner module for log parser module
## authors: luckyducky + lamehacker + calebpitts
################################################################

import csv
import gc
import os
import re
from datetime import datetime
from openpyxl import load_workbook

import pandas as pd

# for holding purposes manually entering machine ID from preferences (needs to be entered as a string)
LogID = []
captureTime = []
bootTime = []
processes = []
disks = []
networkInterfaces = []
battery = []


def addHeaders():
    global LogID, captureTime, bootTime
    global processes, disks
    global networkInterfaces, battery

    LogID.append('Log ID')
    captureTime.append('Log time')
    bootTime.append('bootTime Log')
    processes.append('processes Log')
    disks.append('disks Log')
    networkInterfaces.append('networkInterfaces Log')
    battery.append('battery Log')


def emptyLists():
    global LogID, captureTime, bootTime
    global processes, disks
    global networkInterfaces, battery

    LogID[:] = []
    captureTime[:] = []
    bootTime[:] = []
    processes[:] = []
    disks[:] = []
    networkInterfaces[:] = []
    battery[:] = []

# function to define structure of logID


def generateLogID(logCount, timeForLogID, machineID):
    if logCount <= 9 and logCount >= 0:
        placeholder = '000'
    elif logCount <= 99 and logCount >= 10:
        placeholder = '00'
    elif logCount > 99:
        placeholder = '0'
    date = timeForLogID[13:23]
    LogID.append(machineID + '_' + str(date) + '[' + placeholder + str(logCount) + ']')


def initiator(dataDir, machineID, logPath):
    i = 12
    logCount = 0

    os.chdir(dataDir)
    if os.path.exists('logTrackerLogs.xlsx') == False:
        addHeaders()
    else:
        emptyLists()

    os.chdir(logPath)
    with open('logTracker.log') as logTrackerLog:
        for line in logTrackerLog:

            if line.startswith('<log>') == True:
                i = 1
                continue
            if line.startswith('</log>') == True:
                logCount += 1
                generateLogID(logCount, timeForLogID, machineID)
                i = 11
                continue

            if i == 1:
                captureTime.append(line)
                timeForLogID = line
            elif i == 3:
                bootTime.append(line)
            elif i == 4:
                processes.append(line)
            elif i == 5:
                disks.append(line)
            elif i == 6:
                networkInterfaces.append(line)
            elif i == 7:
                battery.append(line)
            i += 1
    # creates dataframe using list(zip('name of lists to be zipped into dataFrame'))
    df = pd.DataFrame(list(zip(LogID, captureTime, bootTime, processes, disks, networkInterfaces, battery)))

    # creating list of strings for regex to strip prior to putting in dataFrame
    stuffToIgnore = ['<captureTime>', '</captureTime>', '<bootTime>', '</bootTime>', '<processes>', '</processes>',
                     '<disks>', '</disks>', '<networkInterfaces>', '</networkInterfaces>', '<battery>', '</battery>',
                     '\n']

    # needed to add a separate line for \n removal cause it wasn't working in the stuffToIgnore
    df = df.replace("\n", "", regex=True)
    for stuff in stuffToIgnore:
        df = df.replace(stuff, "", regex=True)

    os.chdir(dataDir)

    if os.path.exists('logTrackerLogs.xlsx') == True:
        oldFile = load_workbook('logTrackerLogs.xlsx')
        writer = pd.ExcelWriter('logTrackerLogs.xlsx', engine='openpyxl')
        writer.book = oldFile
        writer.sheets = {ws.title: ws for ws in oldFile.worksheets}
        df.to_excel(writer, sheet_name='logTracker logs', startrow=writer.sheets['logTracker logs'].max_row, index=False,
                    header=False)
        writer.save()
    else:
        writer = pd.ExcelWriter('logTrackerLogs.xlsx')
        df.to_excel(writer, sheet_name='logTracker logs', index=False, header=False)
        writer.save()
