###############################################################################
# Cleaner update for: systemMon.py
# bootClean.py to match format of batteryClean.py
#
# author: LuckyDucky (https://github.com/itsalways5am), lamehacker
#
# Updated: 2018-11-23
###############################################################################
import csv
import gc
import os
import re
from datetime import datetime
from openpyxl import load_workbook

import pandas as pd

# for holding purposes manually entering machine ID from preferences (needs to be entered as a string)
LogID = []
captureTime = []
osFamily = []
osRelease = []
osPlatform = []
osDescription = []
bootTime = []


def addHeaders():
    global LogID, captureTime, supported
    global detected, charge, remaining
    global status, plugged

    LogID.append('Log ID')
    captureTime.append('Log time')
    osFamily.append('OS Family')
    osRelease.append('OS Release')
    osPlatform.append('OS Platform')
    osDescription.append('OS Description')
    bootTime.append('Boot Time')


def emptyLists():
    global LogID, captureTime, osFamily
    global osRelease, osPlatform
    global osDescription, bootTime

    LogID[:] = []
    captureTime[:] = []
    osFamily[:] = []
    osPlatform[:] = []
    osRelease[:] = []
    osDescription[:] = []
    bootTime[:] = []

# function to define structure of logID


def generateLogID(logCount, timeForLogID, machineID):
    if logCount <= 9 and logCount >= 0:
        placeholder = '000'
    elif logCount <= 99 and logCount >= 10:
        placeholder = '00'
    elif logCount > 99:
        placeholder = '0'
    date = timeForLogID[13:23]
    LogID.append(machineID + '_' + str(date) + '[' + placeholder + str(logCount))


def initiator(dataDir, machineID, logPath):
    i = 10
    logCount = 0

    os.chdir(dataDir)
    if os.path.exists('bootTimeLogs.xlsx') == False:
        addHeaders()
    else:
        emptyLists()

    os.chdir(logPath)
    with open('bootTime.log') as bootTimeLog:
        for line in bootTimeLog:

            if line.startswith('<log>') == True:
                i = 1
                continue
            if line.startswith('</log>') == True:
                logCount += 1
                # generateLogID(logCount, timeForLogID, machineID)
                i = 7
                continue

            if i == 1:
                captureTime.append(line)
                timeForLogID = line
                generateLogID(logCount, timeForLogID, machineID)

            elif i == 2:
                osFamily.append(line)
            elif i == 3:
                osRelease.append(line)
            elif i == 4:
                osPlatform.append(line)
            elif i == 5:
                osDescription.append(line)
            elif i == 6:
                bootTime.append(line)

            i += 1

    df = pd.DataFrame(list(zip(LogID, captureTime, osDescription, osFamily, osRelease, osPlatform, bootTime)))

    stuffToIgnore = ['<captureTime>', '</captureTime>', '<osFamily>', '</osFamily>', '<osRelease>', '</osRelease>',
                     '<osPlatform>', '</osPlatform>', '<osDescription>', '</osDescription>', '<bootTime>',
                                     '</bootTime>', '\n']

    for stuff in stuffToIgnore:
        df = df.replace(stuff, "", regex=True)

    os.chdir(dataDir)

    if os.path.exists('bootTimeLogs.xlsx') == True:
        oldFile = load_workbook('bootTimeLogs.xlsx')
        writer = pd.ExcelWriter('bootTimeLogs.xlsx', engine='openpyxl')
        writer.book = oldFile
        writer.sheets = {ws.title: ws for ws in oldFile.worksheets}
        df.to_excel(writer, sheet_name='BootTime logs', startrow=writer.sheets['BootTime logs'].max_row, index=False,
                    header=False)
        writer.save()
    else:
        writer = pd.ExcelWriter('bootTimeLogs.xlsx')
        df.to_excel(writer, sheet_name='BootTime logs', index=False, header=False)
        writer.save()
