###############################################################
# processesClean.py
# last updated 24.11.2018
# purpose : proccesses cleaning module for log parser module
## authors: luckyducky + lamehacker
################################################################

import csv
import gc
import os
import re
from datetime import datetime
from openpyxl import load_workbook
import pandas as pd

# for holding purposes manually entering machine ID from preferences (needs to be entered as a string)
LogID = []
captureTime = []
pid = []
parentPID = []
name = []
cpu = []
priority = []
mem = []
vms = []
rss = []
user = []
path = []


def addHeaders():
    global LogID, captureTime, pid
    global parentPID, name, pctcpu
    global priority, pctmem
    global vms, rss, user, path

    LogID.append('Log ID')
    pid.append('PID')
    parentPID.append('Parent PID')
    name.append('Process name')
    cpu.append('% CPU ')
    priority.append('Priority')
    mem.append('% Mem')
    vms.append('vms')
    rss.append('rss')
    user.append('User')
    path.append('Patj')


def emptyLists():
    global LogID, captureTime, pid
    global parentPID, name, pctcpu
    global priority, pctmem
    global vms, rss, user, path

    LogID[:] = []
    captureTime[:] = []
    pid[:] = []
    parentPID[:] = []
    name[:] = []
    cpu[:] = []
    priority[:] = []
    mem[:] = []
    vms[:] = []
    rss[:] = []
    user[:] = []
    path[:] = []

# function to define structure of logID


def generateLogID(logCount, timeForLogID, machineID):
    if logCount <= 9 and logCount >= 0:
        placeholder = '000'
    elif logCount <= 99 and logCount >= 10:
        placeholder = '00'
    elif logCount > 99:
        placeholder = '0'
    date = timeForLogID[13:23]
    LogID.append(machineID + '_' + str(date) + '[' + placeholder + str(logCount + 1) + ']')


def initiator(dataDir, machineID, logPath):

    logCount = 0
    timestampFix = 0
    timeStamp = ''

    os.chdir(dataDir)
    if os.path.exists('processesLogs.xlsx') == False:
        addHeaders()
    else:
        emptyLists()

    os.chdir(logPath)
    with open('processes.log') as processesLog:
        for line in processesLog:

            if line.startswith('</log>') == True:
                logCount += 1

            if line.startswith('<captureTime>') == True:
                timeStamp = line

            if line.startswith('<pid>') == True:
                pid.append(line)
                captureTime.append(timeStamp)
                timestampFix += 1
                generateLogID(logCount, timeStamp, machineID)

            elif line.startswith('<parentPID>') == True:
                parentPID.append(line)

            elif line.startswith('<name>') == True:
                name.append(line)

            elif line.startswith('<%cpu>') == True:
                cpu.append(line)

            elif line.startswith('<priority>') == True:
                priority.append(line)

            elif line.startswith('<%mem>') == True:
                mem.append(line)

            elif line.startswith('<vms>') == True:
                vms.append(line)

            elif line.startswith('<rss>') == True:
                rss.append(line)

            elif line.startswith('<user>') == True:
                user.append(line)

            elif line.startswith('<path>') == True:
                path.append(line)

    df = pd.DataFrame(
        list(zip(LogID, captureTime, pid, parentPID, name, cpu, priority, mem, vms, rss, user, path)))

    stuffToIgnore = ['<captureTime>', '</captureTime>', '<pid>', '</pid>', '<parentPID>', '</parentPID>', '<name>',
                     '</name>', '<%cpu>', '</%cpu>', '<priority>', '</priority>', '<%mem>', '</%mem>', '<vms>',
                     '</vms>', '<rss>', '</rss>', '<user>', '</user>', '<path>', '</path>', '\n']

    for stuff in stuffToIgnore:
        df = df.replace(stuff, "", regex=True)

    os.chdir(dataDir)

    if os.path.exists('processesLogs.xlsx') == True:
        oldFile = load_workbook('processesLogs.xlsx')
        writer = pd.ExcelWriter('processesLogs.xlsx', engine='openpyxl')
        writer.book = oldFile
        writer.sheets = {ws.title: ws for ws in oldFile.worksheets}
        df.to_excel(writer, sheet_name='Processes logs', startrow=writer.sheets['Processes logs'].max_row, index=False,
                    header=False)
        writer.save()
    else:
        writer = pd.ExcelWriter('processesLogs.xlsx')
        df.to_excel(writer, sheet_name='Processes logs', index=False, header=False)
        writer.save()
