###############################################################
## batteryclean.py
## last updated 23.11.2018 
## purpose : battery cleaner module for log parser module
## authors: luckyducky + lamehacker + calebpitts
################################################################

import csv
import gc
import os
import re
from datetime import datetime
from openpyxl import load_workbook

import pandas as pd

# for holding purposes manually entering machine ID from preferences (needs to be entered as a string)
LogID = []
captureTime = []
supported = []
detected = []
charge = []
remaining = []
status = []
plugged = []

def addHeaders():
    global LogID, captureTime, supported
    global detected, charge, remaining
    global status, plugged

    LogID.append('Log ID')
    captureTime.append('Log time')
    supported.append('Battery supported')
    detected.append('Battery detected')
    charge.append('Current charge')
    remaining.append('Battery reamining for')
    status.append('Charging status')
    plugged.append('Plugged in')

def emptyLists():
    global LogID, captureTime, supported
    global detected, charge, remaining
    global status, plugged

    LogID[:] = []
    captureTime[:] = []
    supported[:] = []
    detected[:] = []
    charge[:] = []
    remaining[:] = []
    status[:] = []
    plugged[:] = []

#function to define structure of logID
def generateLogID(logCount, timeForLogID, machineID):
    if logCount <= 9 and logCount >= 0:
        placeholder = '000'
    elif logCount <= 99 and logCount >= 10:
        placeholder = '00'
    elif logCount > 99:
        placeholder = '0'
    date = timeForLogID[13:23]
    LogID.append(machineID + '_' + str(date) + '[' + placeholder + str(logCount) + ']')

def initiator(dataDir, machineID, logPath):

    i = 10
    logCount = 0
    
    os.chdir(dataDir)
    if os.path.exists('batteryLogs.xlsx') == False:
        addHeaders()
    else:
        emptyLists()

    os.chdir(logPath)
    with open('battery.log') as batteryLog:
        for line in batteryLog:

            if line.startswith('<log>') == True:
                i = 1
                continue

            if line.startswith('</log>') == True:
                logCount += 1
                generateLogID(logCount, timeForLogID, machineID)
                i = 8
                continue

            if i == 1:
                captureTime.append(line)
                timeForLogID = line
            elif i == 2:
                supported.append(line)
            elif i == 3:
                detected.append(line)
            elif i == 4:
                charge.append(line)
            elif i == 5:
                remaining.append(line)
            elif i == 6:
                status.append(line)
            elif i == 7:
                plugged.append(line)
            i += 1

    # creates dataframe using list(zip('name of lists to be zipped into dataFrame'))
    df = pd.DataFrame(list(zip(LogID, captureTime, supported, detected, charge, remaining, status, plugged)))

    # defining items for regex to remove from dataframe
    stuffToIgnore = ['<captureTime>', '</captureTime>',
                    '<supported>', '</supported>',
                    '<detected>', '</detected>',
                    '<charge>', '</charge>',
                    '<status>', '</status>',
                    '<plugged>', '</plugged>',
                    '<remaining>', '</remaining>',
                    '\n']
    
    # regex for replacing stuff with ""
    for stuff in stuffToIgnore:
        df = df.replace(stuff, "", regex=True)

    os.chdir(dataDir)

    if os.path.exists('batteryLogs.xlsx') == True:
        oldFile = load_workbook('batteryLogs.xlsx')
        writer = pd.ExcelWriter('batteryLogs.xlsx', engine='openpyxl')
        writer.book = oldFile
        writer.sheets = {ws.title: ws for ws in oldFile.worksheets}
        df.to_excel(writer,sheet_name='Battery logs', startrow=writer.sheets['Battery logs'].max_row, index = False,header= False)
        writer.save()
    else:
        writer = pd.ExcelWriter('batteryLogs.xlsx')
        df.to_excel(writer, sheet_name='Battery logs', index=False, header=False)
        writer.save()