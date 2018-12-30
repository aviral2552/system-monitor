###############################################################
# disksClean.py
# last updated 23.11.2018
# purpose : disks cleaner module for log parser module
## authors: luckyducky + lamehacker
################################################################

import csv
import gc
import os
import re
from datetime import datetime
from openpyxl import load_workbook

import pandas as pd

# for holding purposes manually entering machine ID from preferences (needs to be entered as a string)
LogID = []
captureTime = []
name = []
size = []
used = []
free = []
pctused = []
fileSys = []
mountPoint = []


def addHeaders():
    # global ( ) allows you to modify the variable outside current scope #makes a global varibale that changes in local context
    global LogID, captureTime, name
    global size, used, free
    global pctused, fileSys, mountPoint

    LogID.append('Log ID')
    captureTime.append('Log time')
    name.append('Disk name')
    size.append('Disk size')
    used.append('Used Memory')
    free.append('Free Memory')
    pctused.append('% Memory Used')
    mountPoint.append('MountPoint')


def emptyLists():
    global LogID, captureTime, name
    global size, used, free
    global pctused, fileSys, mountPoint

    LogID[:] = []
    captureTime[:] = []
    name[:] = []
    size[:] = []
    used[:] = []
    free[:] = []
    pctused[:] = []
    fileSys[:] = []
    mountPoint[:] = []

# function to define structure of logID


def generateLogID(logCount, timeForLogID, machineID):
    if logCount <= 9 and logCount >= 0:
        placeholder = '000'
    elif logCount <= 99 and logCount >= 10:
        placeholder = '00'
    elif logCount > 99:
        placeholder = '0'
    date = timeForLogID[13:23]
    LogID.append(machineID + '_' + str(date) + '[' + placeholder + str(logCount + 1) + ']')


def initiator(dataDir, machineID, logPath):

    logCount = 0  # For holding number of logs
    timestampFix = 0
    timeStamp = ''

    os.chdir(dataDir)
    if os.path.exists('disksLogs.xlsx') == False:
        addHeaders()
    else:
        emptyLists()

    os.chdir(logPath)
    with open('disks.log') as disksLog:
        for line in disksLog:

            if line.startswith('</log>') == True:
                logCount += 1

            if line.startswith('<captureTime>') == True:
                timeStamp = line

            if line.startswith('<name>') == True:
                name.append(line)
                captureTime.append(timeStamp)
                timestampFix += 1
                generateLogID(logCount, timeStamp, machineID)

            elif line.startswith('<size>') == True:
                size.append(line)

            elif line.startswith('<used>') == True:
                used.append(line)

            elif line.startswith('<free>') == True:
                free.append(line)

            elif line.startswith('<%used>') == True:
                pctused.append(line)

            elif line.startswith('<fileSys>') == True:
                fileSys.append(line)

            elif line.startswith('<mountPoint>') == True:
                mountPoint.append(line)

    df = pd.DataFrame(list(zip(LogID, captureTime, name, size, used, free, pctused, fileSys, mountPoint)))

    stuffToIgnore = ['<captureTime>', '</captureTime>', '<name>', '</name>', '<size>', '</size>', '<used>', '</used>',
                     '<free>', '</free>', '<%used>', '</%used>', '<fileSys>', '</fileSys>', '<mountPoint>',
                     '</mountPoint>', '\n']

    for stuff in stuffToIgnore:
        df = df.replace(stuff, "", regex=True)

    os.chdir(dataDir)

    if os.path.exists('disksLogs.xlsx') == True:
        oldFile = load_workbook('disksLogs.xlsx')
        writer = pd.ExcelWriter('disksLogs.xlsx', engine='openpyxl')
        writer.book = oldFile
        writer.sheets = {ws.title: ws for ws in oldFile.worksheets}
        df.to_excel(writer, sheet_name='Disks logs', startrow=writer.sheets['Disks logs'].max_row, index=False,
                    header=False)
        writer.save()
    else:
        writer = pd.ExcelWriter('diskLogs.xlsx')
        df.to_excel(writer, sheet_name='Disk logs', index=False, header=False)
        writer.save()
